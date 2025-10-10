#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from datetime import datetime
import time
import re
import logging
from urllib.parse import urljoin, quote

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/home/ubuntu/mobile_ichiban_scraper.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)

class MobileIchibanScraper:
    def __init__(self):
        self.base_url = "https://www.mobile-ichiban.com"
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        })
    
    def search_by_jan(self, jan_code):
        """JANコードで商品を検索する"""
        try:
            # JANコードで検索
            search_url = f"{self.base_url}/Prod/1"
            params = {
                'search': jan_code
            }
            
            logging.info(f"JANコード {jan_code} で検索中...")
            response = self.session.get(search_url, params=params)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # 商品情報を抽出
            products = self.extract_products_from_page(soup, jan_code)
            
            if products:
                logging.info(f"JANコード {jan_code} で {len(products)} 件の商品が見つかりました")
                return products
            else:
                logging.warning(f"JANコード {jan_code} の商品が見つかりませんでした")
                return []
                
        except Exception as e:
            logging.error(f"JANコード {jan_code} の検索中にエラー: {str(e)}")
            return []
    
    def extract_products_from_page(self, soup, jan_code):
        """ページから商品情報を抽出する"""
        products = []
        
        try:
            # 商品カードを探す
            product_cards = soup.find_all('div', class_='product-card')
            if not product_cards:
                # 別の構造を試す
                product_cards = soup.find_all('div', attrs={'data-jan': True})
            
            for card in product_cards:
                product_info = self.extract_product_info(card, jan_code)
                if product_info:
                    products.append(product_info)
            
            # テキストベースでJANコードを検索
            if not products:
                products = self.extract_products_by_text_search(soup, jan_code)
                
        except Exception as e:
            logging.error(f"商品情報抽出中にエラー: {str(e)}")
        
        return products
    
    def extract_products_by_text_search(self, soup, jan_code):
        """テキストベースでJANコードを含む商品を検索"""
        products = []
        
        try:
            # ページ全体のテキストからJANコードを検索
            page_text = soup.get_text()
            if jan_code in page_text:
                # JANコードが含まれる要素を探す
                jan_elements = soup.find_all(text=re.compile(jan_code))
                
                for element in jan_elements:
                    parent = element.parent
                    while parent:
                        # 価格情報を探す
                        price_text = self.find_price_in_element(parent)
                        if price_text:
                            product_name = self.find_product_name_in_element(parent)
                            
                            product_info = {
                                'jan_code': jan_code,
                                'product_name': product_name or '商品名不明',
                                'price': self.parse_price(price_text),
                                'price_text': price_text,
                                'condition': self.find_condition_in_element(parent),
                                'updated_date': datetime.now().strftime('%Y-%m-%d')
                            }
                            products.append(product_info)
                            break
                        parent = parent.parent
                        
        except Exception as e:
            logging.error(f"テキストベース検索中にエラー: {str(e)}")
        
        return products
    
    def extract_product_info(self, element, jan_code):
        """要素から商品情報を抽出"""
        try:
            product_name = self.find_product_name_in_element(element)
            price_text = self.find_price_in_element(element)
            condition = self.find_condition_in_element(element)
            
            if price_text:
                return {
                    'jan_code': jan_code,
                    'product_name': product_name or '商品名不明',
                    'price': self.parse_price(price_text),
                    'price_text': price_text,
                    'condition': condition or '新品',
                    'updated_date': datetime.now().strftime('%Y-%m-%d')
                }
        except Exception as e:
            logging.error(f"商品情報抽出エラー: {str(e)}")
        
        return None
    
    def find_product_name_in_element(self, element):
        """要素から商品名を探す"""
        try:
            # 商品名を示すクラスやタグを探す
            name_selectors = [
                '.product-name',
                '.item-name',
                'h3',
                'h4',
                '.title',
                'label'
            ]
            
            for selector in name_selectors:
                name_elem = element.select_one(selector)
                if name_elem and name_elem.get_text(strip=True):
                    text = name_elem.get_text(strip=True)
                    if len(text) > 5 and not text.isdigit():  # 商品名らしいテキスト
                        return text
            
            # テキストから商品名らしい部分を抽出
            text = element.get_text(strip=True)
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            for line in lines:
                if any(keyword in line for keyword in ['iPhone', 'iPad', 'Google', 'Galaxy', 'Pixel', 'Redmi']):
                    return line
                    
        except Exception as e:
            logging.error(f"商品名検索エラー: {str(e)}")
        
        return None
    
    def find_price_in_element(self, element):
        """要素から価格を探す"""
        try:
            # 価格パターンを探す
            price_patterns = [
                r'(\d{1,3}(?:,\d{3})*)\s*円',
                r'¥\s*(\d{1,3}(?:,\d{3})*)',
                r'(\d{1,3}(?:,\d{3})*)\s*yen'
            ]
            
            text = element.get_text()
            for pattern in price_patterns:
                matches = re.findall(pattern, text)
                if matches:
                    return matches[0] + '円'
            
            # 価格を示すクラスやタグを探す
            price_selectors = [
                '.price',
                '.amount',
                '.cost',
                'label'
            ]
            
            for selector in price_selectors:
                price_elem = element.select_one(selector)
                if price_elem:
                    price_text = price_elem.get_text(strip=True)
                    if '円' in price_text or '¥' in price_text:
                        return price_text
                        
        except Exception as e:
            logging.error(f"価格検索エラー: {str(e)}")
        
        return None
    
    def find_condition_in_element(self, element):
        """要素からコンディションを探す"""
        try:
            text = element.get_text()
            conditions = ['新品', '中古', '未開封', '開封', 'A級', 'B級', 'C級']
            for condition in conditions:
                if condition in text:
                    return condition
        except Exception:
            pass
        return '新品'
    
    def parse_price(self, price_text):
        """価格テキストから数値を抽出"""
        try:
            # 数字とカンマのみを抽出
            numbers = re.findall(r'[\d,]+', price_text)
            if numbers:
                # カンマを除去して整数に変換
                return int(numbers[0].replace(',', ''))
        except Exception as e:
            logging.error(f"価格パースエラー: {str(e)}")
        return 0

def update_spreadsheet_with_jan_codes(excel_file_path, jan_column='H', price_column='G', date_column='I'):
    """スプレッドシートのJANコードを使って価格を更新する"""
    
    scraper = MobileIchibanScraper()
    
    try:
        # Excelファイルを読み込み
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        
        logging.info(f"スプレッドシート {excel_file_path} を読み込みました")
        
        # ヘッダー行をスキップして処理
        for row_num in range(2, ws.max_row + 1):
            jan_cell = ws[f'{jan_column}{row_num}']
            price_cell = ws[f'{price_column}{row_num}']
            date_cell = ws[f'{date_column}{row_num}']
            
            jan_code = jan_cell.value
            if jan_code and str(jan_code).strip():
                jan_code = str(jan_code).strip()
                logging.info(f"行 {row_num}: JANコード {jan_code} を処理中...")
                
                # JANコードで検索
                products = scraper.search_by_jan(jan_code)
                
                if products:
                    # 最初に見つかった商品の価格を使用
                    product = products[0]
                    price_cell.value = product['price']
                    date_cell.value = product['updated_date']
                    
                    logging.info(f"行 {row_num}: 価格 {product['price']}円 で更新しました")
                else:
                    logging.warning(f"行 {row_num}: JANコード {jan_code} の商品が見つかりませんでした")
                
                # リクエスト間隔を空ける
                time.sleep(2)
        
        # ファイルを保存
        wb.save(excel_file_path)
        logging.info(f"スプレッドシート {excel_file_path} を保存しました")
        
    except Exception as e:
        logging.error(f"スプレッドシート更新エラー: {str(e)}")
        raise

if __name__ == "__main__":
    # テスト実行
    excel_file = "/home/ubuntu/モバイル一番_買取価格一覧.xlsx"
    
    print("JANコード検索システムのテストを開始します...")
    
    # テスト用JANコード
    test_jan_codes = [
        "840353925519",  # Google Pixel 10 Pro
        "840353927292",  # Google Pixel 10 Pro XL
        "4549046144218", # Redmi 12 5G
        "4942857242850", # Galaxy A25 5G
        "4549995560084"  # iPad
    ]
    
    scraper = MobileIchibanScraper()
    
    for jan_code in test_jan_codes:
        print(f"\n--- JANコード {jan_code} のテスト ---")
        products = scraper.search_by_jan(jan_code)
        
        if products:
            for product in products:
                print(f"商品名: {product['product_name']}")
                print(f"価格: {product['price']}円")
                print(f"コンディション: {product['condition']}")
        else:
            print("商品が見つかりませんでした")
        
        time.sleep(1)
    
    print("\nテスト完了")
